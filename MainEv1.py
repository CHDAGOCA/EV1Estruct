from datetime import date, datetime
import re
import csv
import os
import openpyxl
import sqlite3
import sys
import os.path
from sqlite3 import Error
start=False

def obtener_autores():
    cursor.execute("SELECT nombre FROM autores")
    autores = [autor[0] for autor in cursor.fetchall()]
    return autores

def obtener_generos():
    cursor.execute("SELECT nombre FROM generos")
    generos = [genero[0] for genero in cursor.fetchall()]
    return generos

def registro():
    print("-" * 40)
    print("Registro de libro")
    print("-" * 40)
    
    autores = obtener_autores()
    generos = obtener_generos()

    while True:
        while True:
            while True:
                titulo = input("Ingresa el título del libro a registrar: ")
                if titulo.strip() == '':
                    print("El título es un campo obligatorio.")
                else:
                    break

            while True:
                autor = input(f"Ingrese el autor de '{titulo}': ")
                if autor.strip() == '':
                    print("El autor del libro es un campo obligatorio.")
                elif autor not in autores:
                    print("El autor no está registrado. Por favor, elija uno de los autores existentes.")
                    continue
                else:
                    break

            while True:
                genero = input(f"Ingrese el género al que pertenece '{titulo}': ")
                if genero.strip() == '':
                    print("El género del libro es un campo obligatorio.")
                elif genero not in generos:
                    print("El género no está registrado. Por favor, elija uno de los géneros existentes.")
                    continue
                else:
                    break

            while True:
                publicacion = input(f"Ingrese el año de publicación del libro '{titulo}' (YYYY): ")
                if not bool(re.match("^[0-9]{4}$", publicacion)):
                    print("\nEl año de publicación del libro debe tener 4 caracteres numéricos.")
                    continue
                fecha_procesada = datetime.datetime.strptime(publicacion, "%Y").date()
                fecha_actual = datetime.date.today()
                if fecha_procesada > fecha_actual:
                    print("Esta fecha no es válida.")
                else:
                    break

            while True:
                fecha_adquisicion = input("Ingrese la fecha en la que se adquirió el libro (aaaa/mm/dd): ")
                try:
                    fecha_adquisicion = datetime.datetime.strptime(fecha_adquisicion, "%Y/%m/%d").date()
                    dia = fecha_adquisicion.day
                    mes = fecha_adquisicion.month
                    anio = fecha_adquisicion.year
                    break
                except ValueError:
                    print("La fecha ingresada no es válida. Por favor, ingrese una fecha en el formato correcto (aaaa/mm/dd).")
                    continue

            while True:
                isbn = input(f"Ingrese la clave de ISBN del libro '{titulo}': ")
                if len(isbn) == 13 and bool(re.match("^[0-9]{13}$", isbn)):
                    break
                else:
                    print("El ISBN debe tener 13 caracteres numéricos. Vuelva a ingresarlo.")

def registrar_libro(titulo, autor, genero, publicacion, fecha_adquisicion, isbn):
    cursor.execute("INSERT INTO libros (titulo, autor, genero, publicacion, fecha_adquisicion, isbn) VALUES (?, ?, ?, ?, ?, ?)",
                   (titulo, autor, genero, publicacion, fecha_adquisicion, isbn))
    conexion.commit()
    print("El libro se ha registrado exitosamente en la base de datos.")


def consultas():
        while True:
            sub_menu=input("Consultas y Reportajes\n ¿Que accion deseas realizar? \n[1] Consulta por titulo \n[2] Reportajes \n[3] Volver al menú Principal \n")
            sub_menu = sub_menu
            listareport=list(registro_libro.items())
            if sub_menu == "3":
                break
            if sub_menu == "1":
                while True:
                    consulta_titulo=input("Consulta de Titulo \n ¿Que accion deseas realizar? \n[1] Por titulo \n[2] Por ISBN \n[3] Volver al menú de consultas y reportajes \n")
                    consulta_titulo = consulta_titulo
                    if consulta_titulo == "3":
                        break
                    elif consulta_titulo == "1":
                        found=False
                        if listareport:
                            print("Lista de Titulos registrados:")
                            for key,value in listareport:
                                print(value[0])
                        titulo= input("Ingresa el titulo del libro a buscar: ")
                        regisitems=(list(registro_libro.items()))
                        tituloMayus= titulo.upper()
                        for key,value in regisitems:
                            if value[0]==tituloMayus:
                                found=True
                                print("Folio: ",key,"\nTitulo: ",registro_libro[key][0],"\nAutor: ",registro_libro[key][1],"\nGenero: ",registro_libro[key][2],"\nAño de Publicación: ",registro_libro[key][3],"\nFecha_Adquisicion: ",registro_libro[key][4],"\nISBN: ",registro_libro[key][5])
                        if found==False:
                            print("No se encontraron libros con el titulo ", tituloMayus)
                            
                    elif consulta_titulo == "2":
                        found=False
                        isbn= input("Ingresa el isbn del libro a buscar: ")
                        regisitems=(list(registro_libro.items()))
                        for key,value in regisitems:
                            if value[5]==isbn:
                                found=True
                                print("Folio: ",key,"\nTitulo: ",registro_libro[key][0],"\nAutor: ",registro_libro[key][1],"\nGenero: ",registro_libro[key][2],"\nAño de Publicación: ",registro_libro[key][3],"\nFecha_Adquisicion: ",registro_libro[key][4],"\nISBN: ",registro_libro[key][5])
                        if found==False:
                            print("No se encontraron titulos con el ISBN ",isbn)
                    else:
                        print("La opcion ingresada no es correcta, elija de nuevo") 
            if sub_menu == "2":
                while True:
                    reportajes = input("Reportajes \n ¿Que accion deseas realizar? \n[1] Catalogo Completo \n[2] Reportaje por autor \n[3] Reportaje por genero \n[4] Por año de publicacion \n[5] volver al menu de reportaje \n")
                    listareport=list(registro_libro.items())
                    if reportajes=="1":
                        if listareport:
                            print(f"Folio \t Titulo \t\t Autor \t\t\t Genero \t Año_Public \t Fecha_Adq \t ISBN")
                            print("*"*100)
                            for Clave,valor in listareport:
                                print(f'{Clave:4} \t  {valor[0]:15} \t {valor[1]:15} \t {valor[2]:10} \t {valor[3]:5} \t\t {valor[4]:5} \t {valor[5]:15} ')
                            print("*"*100)
                            descarga = input("¿Exportar este reporte? \n[1] Exportar el reportaje en CSV \n[2] Exportar el reportaje en Excel \n[3] No exportar el reporte \n")
                            if descarga=="1":
                                ExportArchComplt_csv()
                            elif descarga=="2":
                                ExportArchComplt_Excel()
                            else:
                                print("Reporte no exportado. \n Volviendo a menu....")
                        else:
                            print("No se han registrado libros.")
                    elif reportajes=="2":
                        foundautor=False
                        if listareport:
                            print("Lista de autores registrados:")
                            for key,value in listareport:
                                print(value[1])
                        autorsel = input("Favor de introducir el autor: ")
                        mayusautorsel=autorsel.upper()
                        if listareport:
                            print(f"Folio \t Titulo \t\t Autor \t\t\t Genero \t Año_Public \t Fecha_Adq \t ISBN")
                            print("*"*100)
                            for key,value in listareport:
                                if value[1]==mayusautorsel:
                                    foundautor=True
                                    print(f'{key:4} \t  {value[0]:15} \t {value[1]:15} \t {value[2]:10} \t {value[3]:5} \t\t {value[4]:5} \t {value[5]:15}')
                            print("*"*100)
                            if foundautor==False:
                                print("No hay libros registrados de este autor")
                                print("*"*100)
                            else:
                                descarga = input("¿Exportar este reporte? \n[1] Exportar el reportaje en CSV \n[2] Exportar el reportaje en Excel \n[3] No exportar el reporte \n")
                                if descarga=="1":
                                    ExportArchAutores_csv(mayusautorsel)
                                elif descarga=="2":
                                    ExportArchAutores_Excel(mayusautorsel)
                                else:
                                    print("Reporte no exportado. \n Volviendo a menu....")
                        else:
                            print("No se han registrado libros por lo que no hay libros de este autor.")
                    elif reportajes=="3":
                        foundgen=False
                        if listareport:
                            print("Lista de generos registrados:")
                            for key,value in listareport:
                                print(value[2])
                        generosel = input("Favor de introducir el genero:   ")
                        mayusgenerosel=generosel.upper()
                        if listareport:
                            print(f"Folio \t Titulo \t\t Autor \t\t\t Genero \t Año_Public \t Fecha_Adq \t ISBN")
                            print("*"*100)
                            for key,value in listareport:
                                if value[2]==mayusgenerosel:
                                    foundgen=True
                                    print(f'{key:4} \t  {value[0]:15} \t {value[1]:15} \t {value[2]:10} \t {value[3]:5} \t\t {value[4]:5} \t {value[5]:15}')
                            print("*"*100)
                            if foundgen==False:
                                print("No hay libros registrados de este genero")
                                print("*"*100)
                            else:
                                descarga = input("¿Exportar este reporte? \n[1] Exportar el reportaje en CSV \n[2] Exportar el reportaje en Excel \n[3] No exportar el reporte \n")
                                if descarga=="1":
                                    ExportArchGenero_csv(mayusgenerosel)
                                elif descarga=="2":
                                    ExportArchGenero_Excel(mayusgenerosel)
                                else:
                                    print("Reporte no exportado. \n Volviendo a menu....")
                        else:
                            print("No se han registrado libros por lo que no hay libros de este genero.")

                    elif reportajes=="4":
                        foundyear=False
                        añosel = input("Favor de introducir el año: ")
                        mayusañosel=añosel.upper()
                        if listareport:
                            print(f"Folio \t Titulo \t\t Autor \t\t\t Genero \t Año_Public \t Fecha_Adq \t ISBN")
                            print("*"*100)
                            for key,value in listareport:
                                if value[3]==mayusañosel:
                                    foundyear=True
                                    print(f'{key:4} \t  {value[0]:15} \t {value[1]:15} \t {value[2]:10} \t {value[3]:5} \t\t {value[4]:5} \t {value[5]:15}')
                            print("*"*100)
                            if foundyear==False:
                                print("No hay libros registrados de este año de publicación")
                                print("*"*100)
                            else:
                                descarga = input("¿Exportar este reporte? \n[1] Exportar el reportaje en CSV \n[2] Exportar el reportaje en Excel \n[3] No exportar el reporte \n")
                                if descarga=="1":
                                    ExportArchAñoPublic_csv(mayusañosel)
                                elif descarga=="2":
                                    ExportArchAñoPublic_Excel(mayusañosel)
                                else:
                                    print("Reporte no exportado. \n Volviendo a menu....")
                        else:
                            print("No se han registrado libros por lo que no hay libros con este año de publicación.")

                    elif reportajes=="5":
                        break
                    else:
                        print("Favor de ingresar una respuesta aceptable.")

            else:
                print("La opcion ingresada no es correcta, elija de nuevo")

def GuardarArchivo():
    listareport=list(registro_libro.items())
    archivo = open("Registro.csv","w",newline="")
    grabador=csv.writer(archivo)
    grabador.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    grabador.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5]) for clave,datos in listareport])
    archivo.close

def CrearTablas():
    file_exists = os.path.exists('Biblioteca.db')
    if not(file_exists):
        try:
            with sqlite3.connect("Biblioteca.db") as conn:
                mi_cursor=conn.cursor()
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS generos (clave INTEGER PRIMARY KEY, GenNombre TEXT NOT NULL);")
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS autores (clave INTEGER PRIMARY KEY, AutNombre TEXT NOT NULL, AutApellidos TEXT NOT NULL);")
                mi_cursor.execute("CREATE TABLE IF NOT EXISTS Libros (clave INTEGER PRIMARY KEY, titulo TEXT NOT NULL, autor INTEGER NOT NULL, \
                                  genero INTEGER NOT NULL, añopublicacion timestamp, ISBN TEXT NOT NULL, \
                                  fechaadq TIMESTAMP, FOREIGN KEY(autor) REFERENCES autores(clave), FOREIGN KEY(genero) REFERENCES genero(clave));")
                print("Tablas creada exitosamente")
        except Error as e:
                print(e)
        except:
                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        finally:
                conn.close()
    else:
        print("Archivo db existente.")

#    try:
        #registro_libro=dict()
        #with open("Registro.csv","r",newline="") as archivo:
            #lector=csv.reader(archivo)
            #next(lector)

            #for clave, titulo,autor,genero,f_publicacion,fecha_adq,isbn in lector:
                #registro_libro[int(clave)]=(titulo,autor,genero,f_publicacion,fecha_adq,isbn)
        #return registro_libro
#    except FileNotFoundError:
        #print("No hay registros previos en la bilbioteca")
        #registro_libro=dict()
        #return registro_libro
#    except csv.Error as fallocsv:
        #print("Ocurrió un error inesperado y no se cargaron los registros")
        #registro_libro=dict()
        #return registro_libro
#    except Exception:
        #print("Deido a un error no se han podido cargar los registros.")
        #registro_libro=dict()
        #return registro_libro

def ExportArchComplt_csv():
    listareport=list(registro_libro.items())
    nombrarch = "ReporteCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5]) for clave,datos in listareport])
    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchAutores_csv(autorsearch):
    listareport=list(registro_libro.items())
    nombrarch = "ReporteAutores" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    for clave,datos in listareport:
        if datos[1]==autorsearch:
            grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5])])
    archivo4.close
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)

def ExportArchGenero_csv(generosearch):
    listareport=list(registro_libro.items())
    nombrarch = "ReporteGenero" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch,"w",newline="")
    grabador1=csv.writer(archivo4)
    grabador1.writerow(("Clave","Titulo","Autor","Genero","f_publicacion","fecha_adquisicion","isbn"))
    for clave,datos in listareport:
        if datos[2]==generosearch:
            grabador1.writerows([(clave,datos[0],datos[1],datos[2],datos[3],datos[4],datos[5])])
    archivo4.close()
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ",nombrarch," y esta en la ruta ",ruta)


def ExportArchAñoPublic_csv(añosearch):
    listareport = list(registro_libro.items())
    nombrarch = "ReporteAñoPublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".csv"
    archivo4 = open(nombrarch, "w", newline="")
    grabador1 = csv.writer(archivo4)
    grabador1.writerow(("Clave", "Titulo", "Autor", "Genero", "f_publicacion", "fecha_adquisicion", "isbn"))
    
    for clave, datos in listareport:
        año = datos[3].split("/")[-1]  
        if año == añosearch:
            grabador1.writerows([(clave, datos[0], datos[1], datos[2], datos[3], datos[4], datos[5])])
    
    archivo4.close()
    ruta = os.getcwd()
    print("El archivo generado tiene por nombre ", nombrarch, " y esta en la ruta ", ruta)


def ExportArchComplt_Excel():
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteCatalogoCompleto" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Catalogo completo"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        hoja.cell(row=i+2, column=2).value = clave
        hoja.cell(row=i+2, column=3).value = valor[0]
        hoja.cell(row=i+2, column=4).value = valor[1]
        hoja.cell(row=i+2, column=5).value = valor[2]
        hoja.cell(row=i+2, column=6).value = valor[3]
        hoja.cell(row=i+2, column=7).value = valor[4]
        hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def ExportArchAutores_Excel(autorsearch):
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteAutores" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Autores"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        if valor[1]==autorsearch:
            hoja.cell(row=i+2, column=2).value = clave
            hoja.cell(row=i+2, column=3).value = valor[0]
            hoja.cell(row=i+2, column=4).value = valor[1]
            hoja.cell(row=i+2, column=5).value = valor[2]
            hoja.cell(row=i+2, column=6).value = valor[3]
            hoja.cell(row=i+2, column=7).value = valor[4]
            hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def ExportArchGenero_Excel(generosearch):
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteGenero" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Genero"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        if valor[2]==generosearch:
            hoja.cell(row=i+2, column=2).value = clave
            hoja.cell(row=i+2, column=3).value = valor[0]
            hoja.cell(row=i+2, column=4).value = valor[1]
            hoja.cell(row=i+2, column=5).value = valor[2]
            hoja.cell(row=i+2, column=6).value = valor[3]
            hoja.cell(row=i+2, column=7).value = valor[4]
            hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)

def ExportArchAñoPublic_Excel(añosearch):
    listareport=list(registro_libro.items())
    ruta = os.getcwd()
    archname = "ReporteAñoPublicacion" + str(datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")) + ".xlsx"
    libro = openpyxl.Workbook()
    libro.iso_dates = True 
    hoja = libro["Sheet"] 
    hoja.title = "Reporte de Año de Publicacion"
    hoja["B1"].value ="Folio"
    hoja["C1"].value ="Titulo"
    hoja["D1"].value ="Autor"
    hoja["E1"].value ="Genero"
    hoja["F1"].value ="Año de Publicación"
    hoja["G1"].value ="Fecha de Adquisición"
    hoja["H1"].value ="ISBN"
    for i, (clave, valor) in enumerate(listareport):
        if valor[3]==añosearch:
            hoja.cell(row=i+2, column=2).value = clave
            hoja.cell(row=i+2, column=3).value = valor[0]
            hoja.cell(row=i+2, column=4).value = valor[1]
            hoja.cell(row=i+2, column=5).value = valor[2]
            hoja.cell(row=i+2, column=6).value = valor[3]
            hoja.cell(row=i+2, column=7).value = valor[4]
            hoja.cell(row=i+2, column=8).value = valor[5]
    libro.save(archname)
    print("El reporte ", archname ," fue creado exitosamente y esta en ",ruta)


def GuardarLibros(titulo,autor,genero,añopub,isbn,fechadqdia,fechadqmes,fechañodq):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (titulo,autor,genero,datetime.datetime(añopub,1,1),isbn,datetime.datetime(fechañodq,fechadqmes,fechadqdia))
            mi_cursor.execute("INSERT INTO Libros (titulo,autor,genero,añopublicacion,ISBN,fechaadq) VALUES(?,?,?,?,?,?)", valores)
        print("Registros agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def GuardarAutores(nombre,apellidos):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (nombre,apellidos)
            mi_cursor.execute("INSERT INTO autores (AutNombre,AutApellidos) VALUES(?,?)", valores)
        print("Autor agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def GuardarGeneros(genero):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor=conn.cursor()
            valores = (genero)
            mi_cursor.execute("INSERT INTO generos (GenNombre) VALUES(?)", valores)
        print("Genero agregado exitosamente.")
    except Error as e:
        print(e)
    except:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()

def registro_autores():
    while True:
        out=False
        autor=""
        apellidos=""
        while True:
            autor=input(f"Favor de ingresar el nombre del autor a registrar ")
            if autor.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",autor))):
                print("\nEl nombre del autor solo puede contener 100 caracteres como máximo entre letras y espacios.")
                continue
            else: 
                break
        while True:
            apellidos=input(f"Favor de ingresar los apellidos del autor a registrar ")
            if apellidos.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",apellidos))):
                print("\nLos apellidos del autor solo pueden contener 100 caracteres como máximo entre letras y espacios.")
                continue
            else: 
                break

        GuardarAutores(autor,apellidos)
        while True:
            salida=input("Introduzca 1 para registrar otro autor\nIntroduzca 2 para salir de la seccion de registro de autores")
            if salida=="1":
                out=False
                break
            elif salida=="2":
                out=True
                break
            else:
                print("Seleccion introducida no valida.")
        if out==True:
            break

def registro_generos():
    while True:
        out=False
        while True:
            genero=input(f"Favor de ingresar el genero a registrar ")
            if genero.strip() == '':
                print("Favor de no dejar el espacio vacio.")
            elif (not bool(re.match("^[A-Za-z ñáéíóúüÑÁÉÍÓÚÜ]{1,100}$",genero))):
                print("\nEl nombre del genero solo puede contener 100 caracteres como máximo entre letras y espacios.")
                continue
            else: 
                break
            
        GuardarGeneros(genero)
        while True:
            salida=input("Introduzca 1 para registrar otro genero\nIntroduzca 2 para salir de la seccion de registro de generos")
            if salida=="1":
                out=False
                break
            elif salida=="2":
                out=True
                break
            else:
                print("Seleccion introducida no valida.")
        if out==True:
            break

def HayAutores():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def HayGeneros():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def HayLibros():
    Existen=False
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Libros ORDER BY clave")
            registros = mi_cursor.fetchall()

        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                Existen=True
        #Si no hay registros en la respuesta
            else:
                Existen=False
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
        return Existen

def ConsultaLibro_TAG(titulo,autor,genero):
    try:
        with sqlite3.connect("Biblioteca.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM Libros ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if titulo==1:
                    print("titulos")
                    print("*" * 30)
                    for claves, titulos, autores, generos,añopub, isbn,fecha in registros:
                        print(f"{titulos:^16}")
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")

            mi_cursor.execute("SELECT * FROM autores ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if autor==1:
                    print("nombre\tapellidos")
                    print("*" * 30)
                    for claves, nombreaut,apellaut in registros:
                        print(f"{nombreaut:^16}\t{apellaut}")
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
            
            mi_cursor.execute("SELECT * FROM generos ORDER BY clave")
            registros = mi_cursor.fetchall()
        #Procedemos a evaluar si hay registros en la respuesta
            if registros:
                if genero==1:
                    print("generos")
                    print("*" * 30)
                    for claves, generonom in registros:
                        print(f"{generonom:^16}")
        #Si no hay registros en la respuesta
            else:
                print("No hay libros registrados.")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        conn.close()
    

CrearTablas()
while True:
    if start==False:
        registro_libro={}
        start=True
    menu_principal=input("Bienvenido a la biblioteca universitaria\n ¿Que accion deseas realizar? \n[1]Registrar un nuevo ejemplar \n[2]Consultas y reportes \n\
                         [3]Registrar un genero\n[4]Registrar un autor\n[5]Salir")
    if menu_principal== "1":
        if HayAutores()==False and HayGeneros()==False:
            print("No hay autores, ni generos registrados por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        elif HayGeneros()==False:
            print("No hay generos registrados, por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        elif HayAutores()==False:
            print("No hay autores registrados, por lo que no se pueden registrar libros.\nVolviendo a menu principal....")
        else:
            registro()
    elif menu_principal=="2":
        if HayLibros()==False:
            print("No hay libros registrados, por lo que no hay libros para consultar o reportar.\nVolviendo a menu principal....")
        else:
            consultas()
    elif menu_principal=="5":
        print("Gracias por visitarnos, vuelva pronto")
        break
    elif menu_principal=="3":
        registro_generos()
    elif menu_principal=="4":
        registro_autores()
    else:
        print("La opcion ingresada no es correcta, elija de nuevo")